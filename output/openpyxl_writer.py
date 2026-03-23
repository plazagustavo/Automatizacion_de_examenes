try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

COMPANY_FIELDS = [
    'Proceso', 'Empresa', 'CUIT', 'Contrato',
    'Domicilio', 'Localidad', 'Provincia', 'Telefono', 'Contacto', 'Email'
]

COLUMN_WIDTHS = [4.86, 53, 13.29]
EXAM_COLUMN_WIDTH = 10.57
HEADER_ROW_HEIGHT = 120


def _thin_border():
    side = Side(style='thin', color='000000')
    return Border(left=side, right=side, top=side, bottom=side)


def write_with_openpyxl(output_file, company_data, final_employees_data, exams_list, exam_count, patient_numbers):
    try:
        wb = Workbook()
        ws = wb.active
        thin_border = _thin_border()

        company_values = {
            'Proceso': '',
            'Empresa': company_data.get('Empresa', ''),
            'CUIT': company_data.get('CUIT', ''),
            'Contrato': company_data.get('Contrato', ''),
            'Domicilio': company_data.get('Domicilio', ''),
            'Localidad': company_data.get('Localidad', ''),
            'Provincia': company_data.get('Provincia', ''),
            'Telefono': company_data.get('Telefono', ''),
            'Contacto': '',
            'Email': company_data.get('Email', ''),
        }

        for i, field in enumerate(COMPANY_FIELDS, start=2):
            ws.cell(row=i, column=2, value=field).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=3, value=company_values[field]).alignment = Alignment(horizontal='left', vertical='center')

        current_row = len(COMPANY_FIELDS) + 3

        headers = ['Id', 'Empleado', 'CUIL'] + exams_list
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            if col_idx > 3:
                cell.alignment = Alignment(text_rotation=90, vertical='bottom', horizontal='center', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='bottom', horizontal='center', wrap_text=True)

        widths = COLUMN_WIDTHS + [EXAM_COLUMN_WIDTH] * len(exams_list)
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

        ws.row_dimensions[current_row].height = HEADER_ROW_HEIGHT

        row_idx = current_row + 1
        last_employee_row = row_idx
        sorted_employees = sorted(final_employees_data.items(), key=lambda x: x[1]['name'].upper())

        for cuil, data in sorted_employees:
            patient_number = patient_numbers[cuil]
            ws.cell(row=row_idx, column=1, value=patient_number).border = thin_border
            ws.cell(row=row_idx, column=2, value=data['name']).border = thin_border
            ws.cell(row=row_idx, column=3, value=data['cuil']).border = thin_border

            for col_idx, exam in enumerate(exams_list, start=4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if any(e.strip() == exam.strip() for e in data['exams']):
                    cell.value = "X"
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            last_employee_row = row_idx
            row_idx += 1

        row_idx += 1
        for exam, count in exam_count.items():
            ws.cell(row=row_idx, column=1, value=count)
            ws.cell(row=row_idx, column=2, value=exam)
            row_idx += 1

        wb.save(output_file)
        return True, last_employee_row

    except Exception as e:
        print(f"Error en openpyxl writer: {e}")
        return False, None
